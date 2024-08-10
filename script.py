# Copyright (c) 2024 by JavierChen(陈佳玮).
# All rights reserved.
#
# Licensed under the MIT License.
# See LICENSE file for more details.
import os
import subprocess
import datetime
import time
import pandas as pd
import pytz
import pysolar.solar as solar
import calculate
from openpyxl import load_workbook
from glob import glob

def validate_longitude(lon):
    return -180 <= lon <= 180

def validate_latitude(lat):
    return -90 <= lat <= 90

def parse_date(date_str):
    if date_str.lower() == 'now':
        return datetime.datetime.now(pytz.timezone('Asia/Shanghai'))
    else:
        try:
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
            return date_obj.replace(tzinfo=pytz.timezone('Asia/Shanghai'))
        except ValueError:
            return None

def getScienceCounts(number):
    return "{:.3E}".format(number)

def format_number(num, precision=6):
    """
    将数字格式化为具有固定小数位数的字符串形式，不足的部分用零填充。

    :param num: 要格式化的数字
    :param precision: 小数位数，默认为6
    :return: 格式化后的字符串
    """
    # 使用f-string进行格式化
    formatted_str = f"{num:.{precision}f}"
    return formatted_str

def day_of_year(dt):
    """
    返回给定日期是一年中的第几天。

    :param dt: date对象
    :return: int, 年积日
    """
    if isinstance(dt, datetime.datetime):
        dt = dt.date()
    # 创建当年1月1日的日期对象
    start_of_year = datetime.date(dt.year, 1, 1)
    delta = dt - start_of_year
    return delta.days + 1


def find_first_excel_file(directory='.'):
    # 查找所有 .xlsx 和 .xls 文件
    files = glob(os.path.join(directory, '*.xlsx')) + glob(os.path.join(directory, '*.xls'))

    # 返回找到的第一个文件，如果没有找到则返回 None
    return files[0] if files else None

def calculate_solar_zenith(longitude, latitude, date=None):
    tz = pytz.timezone('Asia/Shanghai')
    if date is None:
        date = datetime.now(tz).date()

    zenith_angles = {}
    for hour in range(0, 24):
        # 设置时区
        time = datetime.datetime(date.year, date.month, date.day, hour, tzinfo=tz)

        altitude = solar.get_altitude(latitude, longitude, time)
        zenith = 90 - altitude  # Altitude is 90 - zenith

        zenith_angles[hour] = zenith

    return zenith_angles

def tape5Rule(num):
    n = int(num)
    if n == 1:
        return 4
    elif 2 <= n <= 17:
        return 4 + (n - 1)*3

def getTape5(sun_anl, date, factors, Elevation):
    """
    :factors: 影响因素 这里仅仅支持一个列表 里面存储着[pbl_co2_Layer,o3（mg/m3）,co（mg/m3）,so2（mg/m3）,no2（mg/m3）]
    """
    dayOfYear = day_of_year(date)

    with open('tape5.templatetxt', 'r') as template:
        templates = template.readlines()

        temp1 = templates[55]
        templates[55] = temp1.replace('DDD', str(dayOfYear), 1)

        sun_anl = format_number(float(sun_anl), 3)
        temp2 = templates[55]
        templates[55] = temp2.replace('99.999', sun_anl, 1)

        elevation = format_number(Elevation, 3)
        temp3 = templates[55]
        templates[55] = temp3.replace('XXXXX', elevation, 1)

        for i in range(1, 18):
            row = tape5Rule(i)
            if i >= 5:
                factors = [0, 0, 0, 0, 0]
            # Co2
            temp = templates[row]
            co2 = getScienceCounts(factors[0])
            templates[row] = temp.replace('CCCCCCCCC', co2, 1)
            # o3
            temp = templates[row]
            o3 = getScienceCounts(factors[1]) + 'AA2AD2D222DD22'
            templates[row] = temp.replace('DDDDDDDDDAA2AD2D222DD22', o3, 1)
            # co
            temp = templates[row + 1]
            co = format_number(factors[2])
            templates[row + 1] = temp.replace('EEEEEEEE', co, 1)
            # so2
            temp = templates[row + 1]
            so2 = format_number(factors[3])
            templates[row + 1] = temp.replace('FFFFFFFF', so2, 1)
            # no2
            temp = templates[row + 1]
            no2 = format_number(factors[4])
            templates[row + 1] = temp.replace('GGGGGGGG', no2, 1)

        with open('tape5', 'w') as result:
            result.writelines(templates)
        # print('我要睡了')
        # time.sleep(60 * 60)

def geneResult(v1, v2, v3):
    now = datetime.datetime.now()
    folder_name = now.strftime("%Y%m%d")
    result_folder_path = os.path.join("result", folder_name)
    os.makedirs(result_folder_path, exist_ok=True)
    new_file_name = f"{v1}_{v2}_{v3}.tp6"
    source_file_path = os.path.join(".", "tape6")  # 假设 tape6 文件位于当前目录
    destination_file_path = os.path.join(result_folder_path, new_file_name)
    if not os.path.exists(source_file_path):
        print(f"文件 {source_file_path} 不存在，请检查文件路径。")
    else:
        print(f"源文件路径: {source_file_path}")
        print(f"目标文件路径: {destination_file_path}")

        try:
            subprocess.run(f"move {source_file_path} {destination_file_path}",
                           shell=True, check=True)
            print(f"文件已成功创建并重命名为 {new_file_name}。")
        except subprocess.CalledProcessError as e:
            print(f"移动文件时发生错误: {e}")

def main():
    lon = float(input("输入经度: "))
    lat = float(input("输入纬度: "))
    time = input("请输入日期（输入 now 日期为今天）: ")

    if not validate_longitude(lon):
        print("经度无效！请确保在-180到180之间。")
        return

    if not validate_latitude(lat):
        print("纬度无效！请确保在-90到90之间。")
        return

    date = parse_date(time)

    if date is None:
        print("日期格式错误！请输入'now'或'YYYY-MM-DD'格式的日期。")
        return

    zenith_angles = calculate_solar_zenith(lon, lat, date)

    for hour in range(1, 25):
        sun_anl = str(zenith_angles[hour])
        getTape5(sun_anl, date)
        result = subprocess.run('Mod5.2.1.0.exe tap5', stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        # 检查命令是否成功执行
        if result.returncode == 0:
            print("命令成功执行")
        else:
            print(f"命令执行失败，返回码：{result.returncode}")
        date_str = date.strftime('%Y_%m_%d') + "_" +str(hour)
        geneResult(lon, lat, date_str)

def getRes(lon, lat, factors, ele):
    date = parse_date("2020-7-1")
    zenith_angles = calculate_solar_zenith(lon, lat, date)
    for hour in range(0, 24):
        sun_anl = str(zenith_angles[hour])
        if float(sun_anl) < 91.000000:
            getTape5(sun_anl, date, factors, ele)
            result = subprocess.run('Mod5.2.1.0.exe tap5', stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            # 检查命令是否成功执行
            if result.returncode == 0:
                print("命令成功执行")
            else:
                print(f"命令执行失败，返回码：{result.returncode}")
                exit(result.returncode)
            date_str = date.strftime('%Y_%m_%d') + "_" +str(hour)
            geneResult(lon, lat, date_str)

def readXlsx():
    # 加载Excel文件
    df = pd.read_excel('数据(无高度、气压、温度版).xlsxbak')
    factors_guide = ['pbl_co2_Layer（ppmv）', 'o3（mg/m3）', 'co（mg/m3）', 'so2（mg/m3）', 'no2（mg/m3）']
    # 初始化数组来存储提取的数据
    data_arrays = {factor: [] for factor in factors_guide}
    latitudes = []
    longitudes = []

    # 遍历DataFrame中的每一行
    for index, row in df.iterrows():
        # 根据factors_guide中的列名提取数据
        for factor in factors_guide:
            if factor in row.index:
                data_arrays[factor].append(row[factor])

        # 提取经纬度数据
        latitudes.append(row['纬度'])
        longitudes.append(row['经度'])

    # 打印结果
    print("提取的数组数据：", len(data_arrays['pbl_co2_Layer（ppmv）']))
    print("纬度数组：", len(latitudes))
    print("经度数组：", len(longitudes))

# 使用
if __name__ == '__main__':
    start_time = time.time()

    file_path = find_first_excel_file()
    new_path = file_path[2::]
    res_path = new_path[0:1:]

    # 初始化数组来存储提取的数据
    latitudes = []
    longitudes = []
    elevations = []
    results = []
    factors_guide = ['pbl_co2_Layer（ppmv）', 'o3（mg/m3）', 'co（mg/m3）', 'so2（mg/m3）', 'no2（mg/m3）']

    df = pd.read_excel(file_path)

    missing_columns = [col for col in factors_guide if col not in df.columns]
    if missing_columns:
        raise ValueError(f"The following columns are missing from the Excel file: {', '.join(missing_columns)}")

    selected_data = df[factors_guide]

    # 将DataFrame转换为只包含数值的列表
    values_list = selected_data.values.tolist()

    for index, row in df.iterrows():
        latitudes.append(row['纬度'])
        longitudes.append(row['经度'])
        elevations.append(row['高程（km）'])


    for i in range(0, len(longitudes)):
        lon = longitudes[i]
        lat = latitudes[i]
        factors = values_list[i]
        elevation = elevations[i]
        getRes(lon, lat, factors, elevation)
        date_str = datetime.datetime.now(pytz.timezone('Asia/Shanghai')).strftime('%Y%m%d')
        result = calculate.getArea(f"result/{date_str}/" + f"{lon}_{lat}_2020_07_01")
        results.append(result)
        print(result)

    with open(f"{res_path}_result.txt", 'w') as file_res:
        for re in results:
            file_res.write(str(re))
            file_res.write('\n')

    df = pd.read_excel(file_path, engine='openpyxl')
    wb = load_workbook(filename=file_path)
    ws = wb.active

    target_column_name = 'result'
    target_column_index = df.columns.get_loc(target_column_name) + 1 

    last_row = 1  
    for value in results:
        ws.cell(row=last_row + 1, column=target_column_index).value = value
        last_row += 1

    wb.save(f"{res_path}_{res_path}.xlsx")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"代码执行耗时: {elapsed_time:.6f} 秒")